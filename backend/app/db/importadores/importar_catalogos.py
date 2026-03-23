"""
Importador de catálogos documentales desde archivos Excel.

Uso:
    python -m app.db.importadores.importar_catalogos --empresa-id <UUID> [--ruta-base /ruta/a/recursos]

Cada ejecución es idempotente: crea registros si no existen, actualiza si la clave
(etapa_flujo_id + clave normalizada) ya existe.

Estructura esperada de los Excel:
    - Hojas con columnas ÁREA / ETAPA / CATÁLOGO (o similar): se agrupa por ETAPA y se lee el catálogo.
    - Hoja sin fila de encabezados (p. ej. PENAL TRADICIONAL): columnas 1=etapa, 2=documento.
    - Alternativa: filas con texto en MAYÚSCULAS en columna A → encabezado de etapa (fallback).

Pendiente de importar por Excel: CATALOGOS VARIOS ASLIN.xlsx (no está en ARCHIVOS).
"""

import argparse
import logging
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Optional
from uuid import UUID

logger = logging.getLogger("importar_catalogos")
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

# ---------------------------------------------------------------------------
# Configuración de archivos → flujos
# ---------------------------------------------------------------------------

ARCHIVOS: list[dict] = [
    {
        "ruta": "recursos/CATALOGO PENAL-ASLIN.xlsx",
        "hojas": [
            {
                "indice": 0,
                "nombre_flujo": "Litigio Penal - Acusatorio",
                "area_id": UUID("697c62b9-7147-430f-9303-484be5cc49b4"),
            },
            {
                "indice": 1,
                "nombre_flujo": "Litigio Penal - Tradicional",
                "area_id": UUID("697c62b9-7147-430f-9303-484be5cc49b4"),
            },
        ],
    },
    {
        "ruta": "recursos/CATALOGO CIVIL-ASLIN.xlsx",
        "hojas": [
            {
                "indice": 0,
                "nombre_flujo": "Litigio Civil",
                "area_id": UUID("da71ce1c-cc73-4308-be0c-8d70260f7250"),
            },
        ],
    },
    {
        "ruta": "recursos/CATALOGO SERV. PÚB. ASLIN.xlsx",
        "hojas": [
            {
                "indice": 0,
                "nombre_flujo": "Litigio Servidores Públicos",
                "area_id": UUID("0f051097-9fa8-474d-b5fa-8518e070f6fd"),
            },
        ],
    },
]


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def slugify(texto: str) -> str:
    """Convierte texto en clave normalizada (slug)."""
    texto = texto.lower().strip()
    texto = re.sub(r"[áàäâ]", "a", texto)
    texto = re.sub(r"[éèëê]", "e", texto)
    texto = re.sub(r"[íìïî]", "i", texto)
    texto = re.sub(r"[óòöô]", "o", texto)
    texto = re.sub(r"[úùüû]", "u", texto)
    texto = re.sub(r"[ñ]", "n", texto)
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    return texto.strip("_")[:100]


def es_encabezado_etapa(valor: str) -> bool:
    """Heurística para determinar si un valor de celda es un encabezado de etapa.

    Se considera encabezado cuando:
    - Tiene más de 3 caracteres
    - Tiene proporción alta de mayúsculas (> 60%)
    - No contiene caracteres de listas ("•", "-", ".", dígitos iniciales)
    """
    if not valor or len(valor) < 4:
        return False
    texto = valor.strip()
    if texto.startswith(("•", "-", "*")) or re.match(r"^\d+[.)]\s", texto):
        return False
    letras = [c for c in texto if c.isalpha()]
    if not letras:
        return False
    ratio_mayus = sum(1 for c in letras if c.isupper()) / len(letras)
    return ratio_mayus > 0.55


def extraer_bloques(ws) -> list[dict]:
    """Extrae lista de bloques {etapa: str, documentos: [str]} de una hoja Excel."""
    bloques: list[dict] = []
    etapa_actual: Optional[str] = None
    docs_actuales: list[str] = []

    for row in ws.iter_rows(values_only=True):
        # Obtener el primer valor no vacío de la fila
        valor_celdas = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if not valor_celdas:
            continue
        valor_principal = valor_celdas[0]

        if es_encabezado_etapa(valor_principal):
            # Guardar bloque anterior si existe
            if etapa_actual is not None:
                bloques.append({"etapa": etapa_actual, "documentos": docs_actuales})
            etapa_actual = valor_principal.strip().title()  # normalizar capitalización
            docs_actuales = []
        elif etapa_actual is not None and valor_principal:
            # Es un documento dentro de la etapa actual
            docs_actuales.append(valor_principal.strip())

    # Guardar último bloque
    if etapa_actual is not None and (docs_actuales or True):
        bloques.append({"etapa": etapa_actual, "documentos": docs_actuales})

    return bloques


def _norm_cell(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _fila_tiene_header_etapa(fila: tuple) -> bool:
    for c in fila:
        if c is None:
            continue
        t = _norm_cell(c).upper()
        if t == "ETAPA" or t.startswith("ETAPA "):
            return True
    return False


def _indice_columna_catalogo(fila: tuple) -> Optional[int]:
    for i, c in enumerate(fila):
        if c is None:
            continue
        t = _norm_cell(c).upper()
        if "CATAL" in t:
            return i
        raw = _norm_cell(c).upper()
        if "DOCUMENTO" in raw and "CATAL" not in raw:
            return i
    return None


def _agrupar_por_etapa_columnas(
    data_rows: list,
    etapa_idx: int,
    cat_idx: int,
) -> list[dict]:
    bloques_od: OrderedDict[str, list[str]] = OrderedDict()
    last_etapa: Optional[str] = None

    for row in data_rows:
        if not row:
            continue

        def cell(i: int) -> str:
            if i >= len(row) or row[i] is None:
                return ""
            return _norm_cell(row[i])

        et = cell(etapa_idx)
        doc = cell(cat_idx)
        if et:
            last_etapa = et
        if not last_etapa or not doc:
            continue
        etapa_titulo = last_etapa.strip()
        if len(etapa_titulo) > 100:
            etapa_titulo = etapa_titulo[:100]
        if etapa_titulo not in bloques_od:
            bloques_od[etapa_titulo] = []
        bloques_od[etapa_titulo].append(doc)

    return [{"etapa": k, "documentos": v} for k, v in bloques_od.items()]


def extraer_bloques_desde_hoja(ws) -> list[dict]:
    """Igual criterio que el generador SQL de litigio (columna ETAPA + catálogo o 3 columnas)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    primera = rows[0]
    if _fila_tiene_header_etapa(primera):
        etapa_idx = None
        for i, c in enumerate(primera):
            if c is None:
                continue
            u = _norm_cell(c).upper()
            if u == "ETAPA" or u.startswith("ETAPA "):
                etapa_idx = i
                break
        cat_idx = _indice_columna_catalogo(primera)
        if etapa_idx is not None and cat_idx is not None and cat_idx != etapa_idx:
            return _agrupar_por_etapa_columnas(rows[1:], etapa_idx, cat_idx)

    if len(primera) >= 3:
        return _agrupar_por_etapa_columnas(rows, 1, 2)

    return extraer_bloques(ws)


# ---------------------------------------------------------------------------
# Lógica de base de datos
# ---------------------------------------------------------------------------

def get_or_create_flujo(db, empresa_id: UUID, nombre: str, area_id: Optional[UUID] = None):
    """Obtiene o crea un FlujoTrabajo por nombre y empresa."""
    from app.models.flujo_trabajo import FlujoTrabajo

    flujo = (
        db.query(FlujoTrabajo)
        .filter(
            FlujoTrabajo.empresa_id == empresa_id,
            FlujoTrabajo.nombre == nombre,
            FlujoTrabajo.eliminado_en.is_(None),
        )
        .first()
    )
    if flujo:
        if area_id is not None and flujo.area_id != area_id:
            flujo.area_id = area_id
        return flujo, False

    flujo = FlujoTrabajo(
        empresa_id=empresa_id,
        area_id=area_id,
        nombre=nombre,
        descripcion="Importado desde catálogo ASLIN",
        activo=True,
        es_predeterminado=False,
    )
    db.add(flujo)
    db.flush()
    return flujo, True


def get_or_create_etapa(db, flujo_id: UUID, nombre: str, orden: int):
    """Obtiene o crea una EtapaFlujo por nombre y flujo."""
    from app.models.flujo_trabajo import EtapaFlujo

    etapa = (
        db.query(EtapaFlujo)
        .filter(
            EtapaFlujo.flujo_trabajo_id == flujo_id,
            EtapaFlujo.nombre == nombre,
            EtapaFlujo.eliminado_en.is_(None),
        )
        .first()
    )
    if etapa:
        return etapa, False

    etapa = EtapaFlujo(
        flujo_trabajo_id=flujo_id,
        nombre=nombre,
        orden=orden,
        es_obligatoria=True,
        permite_omision=False,
        inhabilita_siguiente=False,
        activo=True,
    )
    db.add(etapa)
    db.flush()
    return etapa, True


def upsert_requisito(db, flujo_id: UUID, etapa_id: UUID, nombre_doc: str, orden: int):
    """Crea o actualiza un EtapaFlujoRequisitoDocumento por (etapa_id, clave)."""
    from app.models.flujo_trabajo import EtapaFlujoRequisitoDocumento

    clave = slugify(nombre_doc)
    requisito = (
        db.query(EtapaFlujoRequisitoDocumento)
        .filter(
            EtapaFlujoRequisitoDocumento.etapa_flujo_id == etapa_id,
            EtapaFlujoRequisitoDocumento.clave == clave,
            EtapaFlujoRequisitoDocumento.eliminado_en.is_(None),
        )
        .first()
    )

    if requisito:
        # Actualizar datos por si cambió el nombre o el orden
        requisito.nombre_documento = nombre_doc
        requisito.orden = orden
        return requisito, False

    requisito = EtapaFlujoRequisitoDocumento(
        flujo_trabajo_id=flujo_id,
        etapa_flujo_id=etapa_id,
        nombre_documento=nombre_doc,
        clave=clave,
        orden=orden,
        es_obligatorio=True,
        permite_upload=True,
        permite_generar=False,
        multiple=False,
        activo=True,
    )
    db.add(requisito)
    return requisito, True


# ---------------------------------------------------------------------------
# Función principal
# ---------------------------------------------------------------------------

def importar(empresa_id: UUID, ruta_base: Path) -> None:
    """Importa todos los catálogos configurados en ARCHIVOS."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl no está instalado. Instálalo con: pip install openpyxl")
        sys.exit(1)

    from app.db.session import SessionLocal

    db = SessionLocal()
    try:
        total_flujos = total_etapas = total_requisitos = 0

        for archivo_cfg in ARCHIVOS:
            ruta_excel = ruta_base / archivo_cfg["ruta"]
            if not ruta_excel.exists():
                logger.warning(f"Archivo no encontrado, omitiendo: {ruta_excel}")
                continue

            logger.info(f"Procesando: {ruta_excel.name}")
            wb = openpyxl.load_workbook(str(ruta_excel), read_only=True, data_only=True)

            for hoja_cfg in archivo_cfg["hojas"]:
                idx = hoja_cfg["indice"]
                nombre_flujo = hoja_cfg["nombre_flujo"]
                area_flujo: Optional[UUID] = hoja_cfg.get("area_id")

                if idx >= len(wb.sheetnames):
                    logger.warning(f"  Hoja índice {idx} no existe en {ruta_excel.name}, omitiendo")
                    continue

                ws = wb.worksheets[idx]
                logger.info(f"  Flujo: {nombre_flujo} (hoja: {ws.title})")

                flujo, creado_flujo = get_or_create_flujo(db, empresa_id, nombre_flujo, area_flujo)
                if creado_flujo:
                    total_flujos += 1
                    logger.info(f"    → Flujo creado: {nombre_flujo}")
                else:
                    logger.info(f"    → Flujo existente: {nombre_flujo}")

                bloques = extraer_bloques_desde_hoja(ws)
                orden_etapa = 1
                for bloque in bloques:
                    etapa_nombre = bloque["etapa"]
                    etapa, creada = get_or_create_etapa(db, flujo.id, etapa_nombre, orden_etapa)
                    if creada:
                        total_etapas += 1
                        logger.info(f"      + Etapa: {etapa_nombre}")
                    orden_etapa += 1

                    orden_req = 1
                    for nombre_doc in bloque["documentos"]:
                        if not nombre_doc:
                            continue
                        _, creado = upsert_requisito(db, flujo.id, etapa.id, nombre_doc, orden_req)
                        if creado:
                            total_requisitos += 1
                        orden_req += 1

            wb.close()

        db.commit()
        logger.info(
            f"\nImportación completada: "
            f"{total_flujos} flujos nuevos, "
            f"{total_etapas} etapas nuevas, "
            f"{total_requisitos} requisitos nuevos."
        )

    except Exception as exc:
        db.rollback()
        logger.exception(f"Error durante la importación: {exc}")
        raise
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importar catálogos documentales ASLIN desde Excel")
    parser.add_argument("--empresa-id", required=True, help="UUID de la empresa destino")
    parser.add_argument(
        "--ruta-base",
        default=".",
        help="Ruta base del proyecto (donde está la carpeta 'recursos/'). Por defecto: directorio actual",
    )
    args = parser.parse_args()

    try:
        empresa_uuid = UUID(args.empresa_id)
    except ValueError:
        logger.error(f"empresa-id inválido: {args.empresa_id}")
        sys.exit(1)

    importar(empresa_uuid, Path(args.ruta_base))
