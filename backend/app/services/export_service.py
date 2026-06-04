"""
Servicio para exportación de datos a Excel, CSV y PDF
"""

import html
import io
import csv
import re
from decimal import Decimal
from typing import List, Dict, Any, Optional

# Importaciones opcionales para evitar errores si no están instaladas
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    Font = None
    Alignment = None
    PatternFill = None
    Border = None
    Side = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    xlsxwriter = None
    XLSXWRITER_AVAILABLE = False


class ExportService:
    """Servicio para exportar datos a diferentes formatos"""

    _MOJIBAKE_MARKERS = re.compile(r"(?:Ã.|Â.|â[€™“”€“˜¢¢])")

    @staticmethod
    def _repair_common_mojibake(value: str) -> str:
        """Repara texto UTF-8 leído como Latin-1/Windows-1252 cuando es evidente."""
        if not ExportService._MOJIBAKE_MARKERS.search(value):
            return value

        candidates = []
        for encoding in ("latin1", "cp1252"):
            try:
                candidates.append(value.encode(encoding).decode("utf-8"))
            except UnicodeError:
                continue

        def mojibake_score(text: str) -> int:
            return len(ExportService._MOJIBAKE_MARKERS.findall(text))

        best = min(candidates or [value], key=mojibake_score)
        return best if mojibake_score(best) < mojibake_score(value) else value

    @staticmethod
    def _scalar_for_export_cell(value: Any) -> Any:
        """Normaliza valores para celda CSV/Excel; decodifica entidades HTML en texto."""
        if isinstance(value, bool):
            return "Sí" if value else "No"
        if value is None:
            return ""
        if isinstance(value, (int, float, Decimal)):
            return value
        return ExportService._repair_common_mojibake(html.unescape(str(value)))

    @staticmethod
    def _label_for_export_header(header: Any) -> str:
        """Etiqueta visible para encabezados de exportación."""
        header_key = str(header)
        if header_key in {"id_normalizado", "id_formato"}:
            return "ID normalizado"
        return header_key.replace('_', ' ').title()

    @staticmethod
    def export_to_excel(
        datos: List[Dict[str, Any]],
        nombre_hoja: str = "Datos",
        titulo: Optional[str] = None,
        columnas: Optional[List[str]] = None,
        columnas_formato_texto: Optional[List[str]] = None,
    ) -> bytes:
        """
        Exporta datos a formato Excel (.xlsx)
        
        Args:
            datos: Lista de diccionarios con los datos
            nombre_hoja: Nombre de la hoja de cálculo
            titulo: Título opcional para el reporte
            columnas: Lista de columnas a incluir (si None, incluye todas)
            columnas_formato_texto: Claves de columnas a forzar como texto en Excel (evita que Excel oculte o malinterprete valores)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Instálelo con: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = nombre_hoja

        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        if not datos:
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        # Determinar columnas
        if columnas:
            headers = columnas
        else:
            headers = list(datos[0].keys())

        texto_cols = set(columnas_formato_texto or [])

        # Agregar título si existe
        row_num = 1
        if titulo:
            ws.merge_cells(f'A{row_num}:{get_column_letter(len(headers))}{row_num}')
            ws[f'A{row_num}'] = titulo
            ws[f'A{row_num}'].font = Font(bold=True, size=14)
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
            row_num += 1

        # Encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = ExportService._label_for_export_header(header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            # Ajustar ancho de columna
            ws.column_dimensions[get_column_letter(col_num)].width = max(len(str(header)) + 2, 15)

        # Datos
        for row_data in datos:
            row_num += 1
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = row_data.get(header, '')
                cell.value = ExportService._scalar_for_export_cell(value)
                if header in texto_cols:
                    cell.number_format = "@"
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Ajustar altura de filas
        for row in ws.iter_rows(min_row=1, max_row=row_num):
            ws.row_dimensions[row[0].row].height = 20

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_to_csv(
        datos: List[Dict[str, Any]],
        columnas: Optional[List[str]] = None,
        delimiter: str = ','
    ) -> str:
        """
        Exporta datos a formato CSV
        
        Args:
            datos: Lista de diccionarios con los datos
            columnas: Lista de columnas a incluir (si None, incluye todas)
            delimiter: Delimitador para el CSV
        """
        if not datos:
            return ""

        output = io.StringIO()
        output.write("\ufeff")
        
        # Determinar columnas
        if columnas:
            headers = columnas
        else:
            headers = list(datos[0].keys())

        writer = csv.DictWriter(output, fieldnames=headers, delimiter=delimiter)
        writer.writeheader()

        for row in datos:
            # Filtrar solo las columnas solicitadas y formatear valores
            filtered_row = {}
            for header in headers:
                value = row.get(header, '')
                v = ExportService._scalar_for_export_cell(value)
                filtered_row[header] = "" if v == "" else str(v)
            writer.writerow(filtered_row)

        return output.getvalue()

    @staticmethod
    def export_to_excel_advanced(
        datos: List[Dict[str, Any]],
        nombre_archivo: str,
        hojas: Optional[Dict[str, List[Dict[str, Any]]]] = None,
        titulo: Optional[str] = None
    ) -> bytes:
        """
        Exporta datos a Excel con múltiples hojas y formato avanzado
        
        Args:
            datos: Datos para la hoja principal
            nombre_archivo: Nombre del archivo
            hojas: Diccionario con nombre_hoja: datos para hojas adicionales
            titulo: Título del reporte
        """
        if not XLSXWRITER_AVAILABLE:
            raise ImportError("xlsxwriter no está instalado. Instálelo con: pip install xlsxwriter")
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        
        # Formato para encabezados
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#366092',
            'font_color': '#FFFFFF',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        # Formato para datos
        data_format = workbook.add_format({
            'border': 1,
            'valign': 'top'
        })

        # Hoja principal
        worksheet = workbook.add_worksheet('Datos')
        
        if datos:
            headers = list(datos[0].keys())
            
            # Escribir título si existe
            row = 0
            if titulo:
                title_format = workbook.add_format({
                    'bold': True,
                    'font_size': 14,
                    'align': 'center'
                })
                worksheet.merge_range(0, 0, 0, len(headers) - 1, titulo, title_format)
                row = 1

            # Escribir encabezados
            for col, header in enumerate(headers):
                worksheet.write(row, col, str(header).replace('_', ' ').title(), header_format)
                worksheet.set_column(col, col, max(len(str(header)) + 2, 15))

            # Escribir datos
            for row_idx, row_data in enumerate(datos, start=row + 1):
                for col_idx, header in enumerate(headers):
                    value = row_data.get(header, '')
                    worksheet.write(
                        row_idx,
                        col_idx,
                        ExportService._scalar_for_export_cell(value),
                        data_format,
                    )

        # Agregar hojas adicionales
        if hojas:
            for nombre_hoja, datos_hoja in hojas.items():
                worksheet = workbook.add_worksheet(nombre_hoja[:31])  # Excel limita a 31 caracteres
                
                if datos_hoja:
                    headers = list(datos_hoja[0].keys())
                    
                    # Encabezados
                    for col, header in enumerate(headers):
                        worksheet.write(0, col, str(header).replace('_', ' ').title(), header_format)
                        worksheet.set_column(col, col, max(len(str(header)) + 2, 15))
                    
                    # Datos
                    for row_idx, row_data in enumerate(datos_hoja, start=1):
                        for col_idx, header in enumerate(headers):
                            value = row_data.get(header, '')
                            worksheet.write(
                                row_idx,
                                col_idx,
                                ExportService._scalar_for_export_cell(value),
                                data_format,
                            )

        workbook.close()
        output.seek(0)
        return output.getvalue()



